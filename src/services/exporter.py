from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from src.db.models import Expense
from src.exceptions import NoDataForExportError


async def build_excel(
    user_id: int,
    year: int,
    month: int,
    session: AsyncSession,
) -> BytesIO:
    date_from = datetime(year, month, 1)
    if month == 12:
        date_to = datetime(year + 1, 1, 1)
    else:
        date_to = datetime(year, month + 1, 1)

    result = await session.execute(
        select(Expense)
        .options(selectinload(Expense.category))
        .where(
            Expense.user_id == user_id,
            Expense.created_at >= date_from,
            Expense.created_at < date_to,
        )
        .order_by(Expense.created_at)
    )
    expenses = list(result.scalars().all())

    if not expenses:
        raise NoDataForExportError(year=year, month=month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    headers = ["Дата", "Категорія", "Сума", "Опис"]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    for expense in expenses:
        cat_label = f"{expense.category.emoji} {expense.category.name}"
        ws.append(
            [
                expense.created_at.strftime("%d.%m.%Y"),
                cat_label,
                float(expense.amount),
                expense.description or "",
            ]
        )
        ws.cell(row=ws.max_row, column=3).number_format = "#,##0.00"

    total = sum((e.amount for e in expenses), Decimal("0"))
    summary_row = ws.max_row + 1
    summary_label = ws.cell(row=summary_row, column=1, value="Всього:")
    summary_label.font = bold
    summary_amount = ws.cell(row=summary_row, column=3, value=float(total))
    summary_amount.font = bold
    summary_amount.number_format = "#,##0.00"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
