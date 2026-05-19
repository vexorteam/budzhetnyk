from datetime import datetime
from decimal import Decimal

import openpyxl
import pytest

from src.db.models import Category, Expense, User
from src.exceptions import NoDataForExportError
from src.services.exporter import build_excel


async def _seed_user_and_category(session):
    user = User(telegram_id=999001, username="exporter_test")
    session.add(user)
    await session.flush()
    cat = Category(
        user_id=None,
        name="Їжа",
        emoji="🍔",
        keywords='["кава"]',
        system_code="food",
    )
    session.add(cat)
    await session.flush()
    await session.commit()
    return user, cat


async def _add_expense(session, user_id, cat_id, amount, description, created_at):
    expense = Expense(
        user_id=user_id,
        category_id=cat_id,
        amount=Decimal(str(amount)),
        description=description,
        created_at=created_at,
    )
    session.add(expense)
    await session.commit()
    return expense


async def test_build_excel_raises_when_no_data(db_session):
    user, _ = await _seed_user_and_category(db_session)
    with pytest.raises(NoDataForExportError) as exc_info:
        await build_excel(user.id, 2025, 1, db_session)
    assert exc_info.value.year == 2025
    assert exc_info.value.month == 1


async def test_build_excel_returns_nonempty_bytesio(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 100, "Кава", datetime(2026, 5, 10))
    buf = await build_excel(user.id, 2026, 5, db_session)
    assert buf.read(4) == b"PK\x03\x04"  # xlsx magic bytes


async def test_build_excel_correct_headers(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 50, "Обід", datetime(2026, 5, 1))
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.cell(1, 1).value == "Дата"
    assert ws.cell(1, 2).value == "Категорія"
    assert ws.cell(1, 3).value == "Сума"
    assert ws.cell(1, 4).value == "Опис"


async def test_build_excel_row_count(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 100, "A", datetime(2026, 5, 1))
    await _add_expense(db_session, user.id, cat.id, 200, "B", datetime(2026, 5, 2))
    await _add_expense(db_session, user.id, cat.id, 300, "C", datetime(2026, 5, 3))
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # 1 header + 3 data rows + 1 summary = 5
    assert ws.max_row == 5


async def test_build_excel_date_format(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 75, "Кава", datetime(2026, 5, 15))
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.cell(2, 1).value == "15.05.2026"


async def test_build_excel_summary_row_label(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 150, "Тест", datetime(2026, 5, 5))
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    last_row = ws.max_row
    assert ws.cell(last_row, 1).value == "Всього:"


async def test_build_excel_summary_row_total(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 100, "A", datetime(2026, 5, 1))
    await _add_expense(db_session, user.id, cat.id, 250, "B", datetime(2026, 5, 2))
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    last_row = ws.max_row
    assert ws.cell(last_row, 3).value == pytest.approx(350.0)


async def test_build_excel_sorted_by_date(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(db_session, user.id, cat.id, 30, "Пізній", datetime(2026, 5, 20))
    await _add_expense(db_session, user.id, cat.id, 10, "Ранній", datetime(2026, 5, 1))
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # row 2 should be the earlier expense
    assert ws.cell(2, 1).value == "01.05.2026"
    assert ws.cell(3, 1).value == "20.05.2026"


async def test_build_excel_excludes_other_months(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(
        db_session, user.id, cat.id, 100, "Квітень", datetime(2026, 4, 15)
    )
    await _add_expense(
        db_session, user.id, cat.id, 200, "Травень", datetime(2026, 5, 10)
    )
    await _add_expense(
        db_session, user.id, cat.id, 300, "Червень", datetime(2026, 6, 1)
    )
    buf = await build_excel(user.id, 2026, 5, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # 1 header + 1 data row + 1 summary = 3
    assert ws.max_row == 3
    assert ws.cell(2, 3).value == pytest.approx(200.0)


async def test_build_excel_december_wraps_year(db_session):
    user, cat = await _seed_user_and_category(db_session)
    await _add_expense(
        db_session, user.id, cat.id, 500, "Грудень", datetime(2026, 12, 25)
    )
    buf = await build_excel(user.id, 2026, 12, db_session)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.max_row == 3  # header + 1 data + summary
